#program to read Reviews from Extracted Data

import os
import openpyxl


os.chdir(r"C:\Users\Pratik\Desktop\Project\ExtractedData")
wb=openpyxl.load_workbook('Extraction.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
ws=wb.active
row_count=sheet.max_row
review_list=[]
for i in range(1,row_count):
    review_list.append(sheet.cell(row=i,column=4).value)



if not os.path.exists(r"C:\Users\Pratik\Desktop\Project\ExtractedData\FetchedData\CleanedData"):
    os.makedirs(r"C:\Users\Pratik\Desktop\Project\ExtractedData\FetchedData\CleanedData")

set_path=(r"C:\Users\Pratik\Desktop\Project\ExtractedData\FetchedData\CleanedData")

os.chdir(set_path)
f=open("Extracted_Reviews.txt","w+",encoding='utf-8')
for review in review_list:
    f.write((review)+'\n')
f.close()


