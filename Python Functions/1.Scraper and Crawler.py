from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException
import logging
import openpyxl
import os


#Check for the directory
os.chdir(r"C:\Users\Pratik\Desktop\Project")

#Logger Syntax
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s- %(message)s')

#Initialization of ChromeDriver for Selenium testing
chrome_path = r"C:\Users\Pratik\Desktop\Project\chromedriver.exe"
driver=webdriver.Chrome(chrome_path)
driver.get("https://www.tripadvisor.in/Hotel_Review-g635747-d477590-Reviews-Keys_Ronil_Resort-Baga_Goa.html")
driver.maximize_window()


logging.info('Executing Main Function to scrape the contents')

##################################################################################################################################################
os.chdir(r"C:\Users\Pratik\Desktop\Project\ExtractedData")
def mainfunction():
    users=[]
    dates=[]
    heading=[]
    text=[]

    wb = openpyxl.load_workbook('Extraction.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    ws = wb.active
    
#Hotel Name
    hname = driver.find_element_by_id('HEADING').text


#User Names
    UserReviewsSection = driver.find_element_by_id('REVIEWS')
    Usernames = UserReviewsSection.find_elements_by_class_name('prw_reviews_basic_review')
    for user in Usernames:
        users.append(user.find_element_by_class_name('expand_inline').text)
	

#Dates
    datediv = driver.find_elements_by_css_selector('div > div.col2of2 > div > div.wrap > div.rating.reviewItemInline > span.ratingDate.relativeDate')
    for date in datediv:
        dates.append(date.get_attribute("title"))


#Review Title
    ReviewSection = driver.find_element_by_id('REVIEWS')
    TitleHeader = ReviewSection.find_elements_by_class_name('prw_reviews_basic_review')
    for Titles in TitleHeader:
        heading.append(Titles.find_element_by_class_name('noQuotes').text)


#Reviews 
    linkdiv = driver.find_element_by_class_name('expandLink')
    linkspan = linkdiv.find_element_by_class_name('ulBlueLinks')
    linkspan.click()

    try:
        WebDriverWait(driver,10).until(ec.presence_of_element_located((By.CLASS_NAME,"no_padding")))
        close1 = driver.find_element_by_css_selector('body > div> span > div.ui_close_x')
        close1.click()

    except TimeoutException:
        logging.warning('Exceeded explicit wait for popup!')


    reviews = driver.find_elements_by_css_selector(' div > div.col2of2 > div > div.wrap > div > div > p')
    for review in reviews:
        text.append(review.text)

    
    logging.info('Writing contents to the excel file')
#push the contents to the dictionary
    
    #Essential to open everytime
    wb = openpyxl.load_workbook('Extraction.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    ws = wb.active
    
    #Load the newest max_row value
    maxentry=ws.max_row

    logging.info('Writing Users')
    trigger=False
    temp=maxentry
    for row, i in enumerate(users):
        if not trigger:
            column_cell = 'A'
            ws[column_cell+str(temp)] = str(i)
            #print(str(i))
            temp+=1
        
    logging.info('Writing Dates')
    temp=maxentry
    for row, i in enumerate(dates):
        column_cell = 'B'
        ws[column_cell+str(temp)] = str(i)
        #print(str(i))
        temp+=1
    
    logging.info('Writing Titles')
    temp=maxentry
    for row, i in enumerate(heading):
        column_cell = 'C'
        ws[column_cell+str(temp)] = str(i)
        #print(str(i))
        temp+=1
        
    logging.info('Writing Reviews')    
    temp=maxentry
    for row, i in enumerate(text):
        column_cell = 'D'
        ws[column_cell+str(temp)] = str(i)
        #print(str(i))
        temp+=1

    logging.info('Finishing writing to excel file')    
    wb.save("Extraction.xlsx")

    
    #Move to next page
    logging.info('Navigating to the next page')
    nextpage()

##################################################################################################################################################


#To follow successive pages and scrape the content
    
def nextpage():
    nextpage = driver.find_element_by_css_selector('#REVIEWS > div.deckTools.btm.test > div > a.nav.next.rndBtn.ui_button.primary.taLnk').click()
    try:
        WebDriverWait(driver,10).until(ec.presence_of_element_located((By.XPATH,'//*[@id="overlayRegFrame"]')))
        close2 = driver.find_element_by_xpath('/html/body/div[9]/span/div[2]').click()
    
        WebDriverWait(driver,10).until(ec.presence_of_element_located((By.CLASS_NAME,"RBP_PCBSurvey_All_EN")))
        close3 = driver.find_element_by_class_name('ui_close_x')
        close3.click()
   
    except TimeoutException:
        logging.warning('Exceeded explicit wait for popup!')

    logging.info('Entering Main Function from within nextpage()')
    mainfunction()




#------------------------------------------------------------------------------------------------------------------------------------------------------#
    
logging.info('Start of program')
mainfunction()
